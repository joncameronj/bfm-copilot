#!/usr/bin/env python3
"""
Convert BFM Master Protocol Key XLSX to structured markdown files.

Reads each sheet from the master protocol XLSX and generates a well-structured
markdown file optimized for semantic chunking and RAG retrieval.

Format: section-per-row (not tables) so each protocol entry becomes a
self-contained chunk with full context.

Usage:
    python scripts/convert_master_xlsx.py [--input PATH] [--output-dir PATH]
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


DEFAULT_INPUT = Path(__file__).parent.parent.parent / "BFM_Frequency_Mapping_Master.xlsx"
DOWNLOADS_INPUT = Path.home() / "Downloads" / "BFM Master Protocol Key (2).xlsx"
DEFAULT_OUTPUT = Path(__file__).parent.parent.parent / "agent-assets" / "master-protocols"


def safe_str(value) -> str:
    """Convert cell value to string, handling None."""
    if value is None:
        return ""
    return str(value).strip()


def is_section_header(text: str) -> tuple[bool, str]:
    """
    Detect if a cell value is a section header (e.g., '── LIVER ──', '— THYROID —').

    Returns:
        Tuple of (is_header, clean_section_name)
    """
    stripped = text.strip()
    # Check for various dash/line patterns used as section separators
    # Handles em-dash (—), en-dash (–), box-drawing (─, ─), hyphens (--)
    clean = stripped.strip("—–─-─ \t")
    if not clean:
        return False, ""
    # If the original had leading/trailing dashes and remaining text is mostly uppercase
    has_dashes = any(c in stripped for c in "—–─-")
    if has_dashes and (clean.isupper() or clean.upper() == clean):
        return True, clean
    return False, ""


def convert_deal_breakers(ws) -> str:
    """Convert Deal Breakers sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Deal Breakers",
        "",
        "These are the 7 critical deal breakers that MUST be addressed before proceeding",
        "with any frequency protocol. If a patient presents with any of these findings,",
        "the corresponding protocol takes absolute priority.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    # Find header row (first row with content)
    header_idx = 0
    for i, row in enumerate(rows):
        if row[0] and "deal breaker" in str(row[0]).lower():
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        deal_breaker = vals[0]
        lines.append(f"## Deal Breaker: {deal_breaker}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_hrv_brainwave(ws) -> str:
    """Convert HRV & Brainwave Mapping sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: HRV & Brainwave Mapping",
        "",
        "Maps HRV autonomic patterns and brainwave findings to frequency protocols.",
        "Use this to determine the correct frequency based on HRV or D-Pulse brainwave results.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        if row[0] and ("hrv" in str(row[0]).lower() or "finding" in str(row[0]).lower()):
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    current_section = None
    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        # Detect section headers (e.g., "AUTONOMIC PATTERNS", "BRAINWAVE PATTERNS")
        is_header, section_name = is_section_header(vals[0])
        if is_header:
            current_section = section_name
            lines.append(f"## {current_section}")
            lines.append("")
            continue

        # Regular data row
        finding = vals[0]
        lines.append(f"### {finding}")
        lines.append("")

        if current_section:
            lines.append(f"**Section:** {current_section}")
            lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_dpulse_organ(ws) -> str:
    """Convert D-Pulse Organ Mapping sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: D-Pulse Organ Mapping",
        "",
        "Maps low organ scores from D-Pulse testing to frequency protocols.",
        "When a D-Pulse shows low scores for specific organs, use this reference",
        "to determine the correct frequency protocol and supplementation.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        if row[0] and ("d-pulse" in str(row[0]).lower() or "finding" in str(row[0]).lower()):
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        finding = vals[0]
        lines.append(f"## D-Pulse Finding: {finding}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_lab_diagnostic(ws) -> str:
    """Convert Lab & Diagnostic Mapping sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Lab & Diagnostic Mapping",
        "",
        "Maps lab markers and diagnostic findings to frequency protocols and supplementation.",
        "This is the primary reference for translating blood work, urinalysis, and other",
        "diagnostic results into specific frequency protocol recommendations.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        if row[0] and ("lab" in str(row[0]).lower() or "marker" in str(row[0]).lower()):
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    current_section = None
    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        is_header, section_name = is_section_header(vals[0])
        if is_header:
            current_section = section_name
            lines.append(f"## Lab Category: {current_section}")
            lines.append("")
            continue

        # Regular data row
        marker = vals[0]
        lines.append(f"### Lab Marker: {marker}")
        lines.append("")

        if current_section:
            lines.append(f"**Category:** {current_section}")
            lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_condition_protocols(ws) -> str:
    """Convert Condition Protocols sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Condition-Specific Protocols",
        "",
        "Disease and condition-specific protocol stacks with frequency combinations,",
        "supplementation, and treatment timelines. Organized by care category",
        "(Thyroid, Neurological, Hormones, Diabetes).",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        if row[0] and ("condition" in str(row[0]).lower() or "protocol" in str(row[0]).lower()):
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    current_section = None
    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        is_header, section_name = is_section_header(vals[0])
        if is_header:
            current_section = section_name
            lines.append(f"## Care Category: {current_section}")
            lines.append("")
            continue

        condition = vals[0]
        lines.append(f"### Condition: {condition}")
        lines.append("")

        if current_section:
            lines.append(f"**Care Category:** {current_section}")
            lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_five_levers(ws) -> str:
    """Convert Five Levers sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: The Five Levers",
        "",
        "The 5 master levers that govern metabolic health in the BFM framework.",
        "These are the foundational markers that must be assessed and optimized:",
        "Melatonin, Leptin, MSH, Vitamin D, and UB Rates.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        first = safe_str(row[0]).lower() if row[0] else ""
        if "lever" in first or "master" in first:
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        lever = vals[0]
        lines.append(f"## Master Lever: {lever}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_supplement_reference(ws) -> str:
    """Convert Supplement Reference sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Supplement Reference",
        "",
        "Complete supplement catalog with dosages, timing, indications, and brands.",
        "This is the authoritative reference for all supplementation recommendations",
        "in the BFM protocol system.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        first = safe_str(row[0]).lower() if row[0] else ""
        if "supplement" in first or "name" in first:
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        supplement = vals[0]
        lines.append(f"## Supplement: {supplement}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_mitochondrial(ws) -> str:
    """Convert Mitochondrial Frequencies sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Mitochondrial Frequencies",
        "",
        "All MIT/Mito frequency settings with mechanisms and indications.",
        "These frequencies target mitochondrial function and cellular energy production.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        first = safe_str(row[0]).lower() if row[0] else ""
        if "mito" in first or "frequency" in first or "setting" in first:
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        freq = vals[0]
        lines.append(f"## Mitochondrial Frequency: {freq}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def convert_contraindications(ws) -> str:
    """Convert Contraindications sheet to markdown."""
    lines = [
        "# BFM Master Protocol Key: Contraindications & Safety Warnings",
        "",
        "CRITICAL SAFETY INFORMATION: These contraindications must be checked before",
        "ANY frequency protocol is recommended. Failure to screen for these conditions",
        "could result in adverse events.",
        "",
        "---",
        "",
    ]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "\n".join(lines)

    header_idx = 0
    for i, row in enumerate(rows):
        first = safe_str(row[0]).lower() if row[0] else ""
        if "contra" in first or "condition" in first or "warning" in first:
            header_idx = i
            break

    headers = [safe_str(h) for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    for row in data_rows:
        vals = [safe_str(v) for v in row]
        if not vals[0]:
            continue

        condition = vals[0]
        lines.append(f"## Contraindication: {condition}")
        lines.append("")

        for i, header in enumerate(headers[1:], start=1):
            if i < len(vals) and vals[i]:
                lines.append(f"**{header}:** {vals[i]}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


# Map sheet names to converter functions
SHEET_CONVERTERS = {
    "Deal Breakers": ("01-deal-breakers.md", convert_deal_breakers),
    "HRV & Brainwave Mapping": ("02-hrv-brainwave-mapping.md", convert_hrv_brainwave),
    "D-Pulse Organ Mapping": ("03-dpulse-organ-mapping.md", convert_dpulse_organ),
    "Lab & Diagnostic Mapping": ("04-lab-diagnostic-mapping.md", convert_lab_diagnostic),
    "Condition Protocols": ("05-condition-protocols.md", convert_condition_protocols),
    "Five Levers": ("06-five-levers.md", convert_five_levers),
    "Supplement Reference": ("07-supplement-reference.md", convert_supplement_reference),
    "Mitochondrial Frequencies": ("08-mitochondrial-frequencies.md", convert_mitochondrial),
    "Contraindications": ("09-contraindications.md", convert_contraindications),
}


def convert_workbook(input_path: Path, output_dir: Path) -> dict:
    """
    Convert entire workbook to structured markdown files.

    Returns:
        Dict with conversion statistics
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    stats = {"converted": 0, "skipped": 0, "errors": []}

    print(f"Input:  {input_path}")
    print(f"Output: {output_dir}")
    print(f"Sheets found: {wb.sheetnames}")
    print("=" * 60)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CONVERTERS:
            print(f"  SKIP: '{sheet_name}' (no converter defined)")
            stats["skipped"] += 1
            continue

        filename, converter = SHEET_CONVERTERS[sheet_name]
        ws = wb[sheet_name]

        try:
            markdown = converter(ws)
            output_path = output_dir / filename
            output_path.write_text(markdown, encoding="utf-8")

            line_count = markdown.count("\n")
            print(f"  OK:   '{sheet_name}' -> {filename} ({line_count} lines)")
            stats["converted"] += 1

        except Exception as e:
            print(f"  ERR:  '{sheet_name}': {e}")
            stats["errors"].append({"sheet": sheet_name, "error": str(e)})

    print("=" * 60)
    print(f"Converted: {stats['converted']}, Skipped: {stats['skipped']}, Errors: {len(stats['errors'])}")

    return stats


def main():
    parser = argparse.ArgumentParser(description="Convert BFM Master Protocol XLSX to markdown")
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Path to input XLSX file",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Output directory for markdown files",
    )
    args = parser.parse_args()

    # Determine input file
    input_path = args.input
    if input_path is None:
        # Prefer Downloads version (has correct branded names)
        if DOWNLOADS_INPUT.exists():
            input_path = DOWNLOADS_INPUT
            print(f"Using Downloads version (branded names): {input_path}")
        elif DEFAULT_INPUT.exists():
            input_path = DEFAULT_INPUT
            print(f"Using repo version: {input_path}")
        else:
            print("Error: No input XLSX found. Provide --input PATH")
            sys.exit(1)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        sys.exit(1)

    convert_workbook(input_path, args.output_dir)


if __name__ == "__main__":
    main()
